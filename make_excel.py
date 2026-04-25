import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

valid = []
with open('output.jsonl') as f:
    for line in f:
        try:
            valid.append(json.loads(line))
        except:
            pass

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "REVISE Results"

headers = ["#", "Question", "Original Claim", "Revised Claim"]
col_widths = [5, 35, 60, 60]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="1F4E79")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w

for i, obj in enumerate(valid, start=1):
    row = i + 1
    values = [i, obj.get("question_text",""), obj.get("claim",""), obj.get("revised_claim","")]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 100

wb.save("output_partial.xlsx")
print(f"Saved {len(valid)} claims to output_partial.xlsx")
