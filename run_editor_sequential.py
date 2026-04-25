import argparse
import json
import logging
import os
import warnings
warnings.filterwarnings("ignore", message="Can't initialize NVML")

import jsonlines
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import nltk
from tqdm import tqdm

from prompts import hallucination_prompts, rarr_prompts
from utils import (
    agreement_gate,
    editor,
    evidence_selection,
    hallucination,
    search,
    question_generation,
)

nltk.download('punkt', quiet=True)
nltk.download('punkt_tab', quiet=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("run_editor_sequential.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def get_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Revise misinformation claims from a JSONL file using the REVISE pipeline."
    )
    parser.add_argument("--input_file",  type=str, required=True,  help="Input JSONLines file of claims.")
    parser.add_argument("--output_file", type=str, required=True,  help="Output JSONL file.")
    parser.add_argument("--model",       type=str, default="gpt-3.5-turbo", help="OpenAI model.")
    parser.add_argument("--temperature_qgen", type=float, default=0.7, help="Temperature for question generation.")
    parser.add_argument("--hallucinate_evidence", action="store_true",
                        help="Generate evidence via LLM instead of web retrieval.")
    parser.add_argument("--max_search_results_per_query",    type=int, default=3)
    parser.add_argument("--max_sentences_per_passage",       type=int, default=4)
    parser.add_argument("--sliding_distance",                type=int, default=1)
    parser.add_argument("--max_passages_per_search_result",  type=int, default=1)
    return parser.parse_args()


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------
def read_claims(input_file: str):
    """
    Read claims from a JSON or JSONL file.
    Supports:
      - JSONL: one JSON object per line, each with a 'claim' field
      - JSON array: a list of objects, each with a 'claim' field
      - JSON array of strings: each string becomes a claim
      - JSON object with a 'data' or 'claims' key containing a list
    """
    claims = []
    try:
        # First try as regular JSON
        with open(input_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            for item in data:
                if isinstance(item, str):
                    claims.append({"claim": item})
                elif isinstance(item, dict):
                    # Support 'claim', 'text', 'passage', or 'long_answer' as the claim field
                    claim_text = (
                        item.get("long_answer")
                        or item.get("REVISE")
                        or item.get("claim")
                        or item.get("text")
                        or item.get("passage")
                        or item.get("long_answer")
                        or ""
                    )
                    claims.append({**item, "claim": claim_text})
        elif isinstance(data, dict):
            # Try common wrapper keys
            for key in ("data", "claims", "examples", "items"):
                if key in data and isinstance(data[key], list):
                    return read_claims.__wrapped__(data[key])
            # Single object
            claims.append(data)

        logger.info(f"Loaded {len(claims)} claims from {input_file} (JSON format)")

    except json.JSONDecodeError:
        # Fall back to JSONL
        try:
            with jsonlines.open(input_file) as reader:
                for obj in reader:
                    if isinstance(obj, str):
                        claims.append({"claim": obj})
                    else:
                        claim_text = (
                            obj.get("claim")
                            or obj.get("text")
                            or obj.get("passage")
                            or obj.get("long_answer")
                            or obj.get("misinformation_passage")
                            or ""
                        )
                        claims.append({**obj, "claim": claim_text})
            logger.info(f"Loaded {len(claims)} claims from {input_file} (JSONL format)")
        except Exception as e:
            logger.error(f"Failed to read input file: {e}")
            raise
    except Exception as e:
        logger.error(f"Failed to read input file: {e}")
        raise

    return claims


def write_result(writer: jsonlines.Writer, result: dict) -> None:
    try:
        writer.write(result)
    except Exception as e:
        logger.error(f"Failed to write result: {e}")


# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------
def split_into_atomic_statements(text: str):
    """Split text into sentences using nltk with fallback to regex."""
    text = text.strip()
    if not text:
        return []
    try:
        # Try punkt_tab first (newer NLTK), then punkt
        try:
            from nltk.tokenize import PunktSentenceTokenizer
            tokenizer = nltk.data.load("tokenizers/punkt_tab/english.pickle")
            return tokenizer.tokenize(text)
        except Exception:
            return nltk.sent_tokenize(text)
    except Exception:
        # Fallback: simple regex sentence splitter
        import re
        sentences = re.split(r'(?<=[.!?])\s+', text)
        return [s.strip() for s in sentences if s.strip()]


def process_atomic_statement(
    statement: str,
    accumulated_context: str,
    model: str,
    temperature_qgen: float,
    search_params: dict,
    hallucinate_evidence: bool = False,
) -> dict:
    """
    For one atomic statement:
      1. Generate fact-checking questions.
      2. Retrieve evidence (web or hallucinated).
      3. Run agreement gate + editor for each evidence item.
    Returns a dict with the original and revised statement plus metadata.
    """
    try:
        # ---- Question generation ------------------------------------------
        questions = question_generation.run_rarr_question_generation(
            claim=statement,
            model=model,
            prompt=(
                rarr_prompts.CONTEXTUAL_QGEN_PROMPT
                if accumulated_context
                else rarr_prompts.QGEN_PROMPT
            ),
            temperature=temperature_qgen,
            context=accumulated_context,
            required_questions=3,
        )

        all_evidences = []
        revised_statement = statement

        for question in questions:
            # ---- Evidence retrieval ---------------------------------------
            if hallucinate_evidence:
                ev_result = hallucination.run_evidence_hallucination(
                    query=question,
                    model=model,
                    prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
                    context=accumulated_context if accumulated_context else None,
                    atomic_statement=statement,
                )
                evidences = [ev_result] if ev_result.get("text") else []
            else:
                evidences = search.run_search(
                    query=question,
                    context=accumulated_context if accumulated_context else None,
                    atomic_statement=statement,
                    **search_params,
                )

            # ---- Agreement gate + editor ----------------------------------
            for evidence_item in evidences:
                ev_text = evidence_item.get("text", "")
                if not ev_text:
                    continue

                gate_result = agreement_gate.run_agreement_gate(
                    claim=revised_statement,
                    query=question,
                    evidence=ev_text,
                    model=model,
                    prompt=(
                        rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT
                        if accumulated_context
                        else rarr_prompts.AGREEMENT_GATE_PROMPT
                    ),
                    context=accumulated_context if accumulated_context else None,
                )
                all_evidences.append({
                    "text": ev_text,
                    "query": question,
                    "gate_result": gate_result,
                })

                if gate_result.get("is_open"):
                    edited = editor.run_rarr_editor(
                        claim=revised_statement,
                        query=question,
                        evidence=ev_text,
                        model=model,
                        prompt=(
                            rarr_prompts.CONTEXTUAL_EDITOR_PROMPT
                            if accumulated_context
                            else rarr_prompts.EDITOR_PROMPT
                        ),
                        context=accumulated_context if accumulated_context else None,
                    )
                    if edited.get("text"):
                        revised_statement = edited["text"]

        return {
            "original_statement": statement,
            "revised_statement": revised_statement,
            "questions": questions,
            "evidences": all_evidences,
        }

    except Exception as e:
        logger.error(f"Error processing statement '{statement[:60]}': {e}")
        return {
            "original_statement": statement,
            "revised_statement": statement,
            "questions": [],
            "evidences": [],
        }


def revise_claim(
    claim_obj: dict,
    model: str,
    temperature_qgen: float,
    search_params: dict,
    hallucinate_evidence: bool,
) -> dict:
    """
    Revise all atomic statements in one claim object.
    Input JSON object must have a 'claim' key (string).
    """
    original_text = claim_obj.get("claim", "")
    if not original_text.strip():
        logger.warning("Empty claim encountered, skipping.")
        return {**claim_obj, "revised_claim": original_text, "statement_results": []}

    statements = split_into_atomic_statements(original_text)
    revised_statements = []
    statement_results = []
    accumulated_context = ""

    for statement in statements:
        result = process_atomic_statement(
            statement=statement,
            accumulated_context=accumulated_context,
            model=model,
            temperature_qgen=temperature_qgen,
            search_params=search_params,
            hallucinate_evidence=hallucinate_evidence,
        )
        revised_statements.append(result["revised_statement"])
        statement_results.append(result)
        # Update context with revised text so far
        accumulated_context = " ".join(revised_statements)

    revised_text = " ".join(revised_statements)

    return {
        **claim_obj,
        "revised_claim": revised_text,
        "statement_results": statement_results,
    }


def revise_and_save_jsonl(
    claims,
    output_jsonl: str,
    model: str,
    temperature_qgen: float,
    max_search_results_per_query: int,
    max_sentences_per_passage: int,
    sliding_distance: int,
    max_passages_per_search_result: int,
    hallucinate_evidence: bool,
) -> None:
    search_params = {
        "max_search_results_per_query": max_search_results_per_query,
        "max_sentences_per_passage": max_sentences_per_passage,
        "sliding_distance": sliding_distance,
        "max_passages_per_search_result_to_return": max_passages_per_search_result,
    }

    os.makedirs(os.path.dirname(os.path.abspath(output_jsonl)), exist_ok=True)

    with jsonlines.open(output_jsonl, mode="w") as writer:
        for claim_obj in tqdm(claims, desc="Revising claims"):
            result = revise_claim(
                claim_obj=claim_obj,
                model=model,
                temperature_qgen=temperature_qgen,
                search_params=search_params,
                hallucinate_evidence=hallucinate_evidence,
            )
            write_result(writer, result)
            logger.info(f"Revised: {claim_obj.get('claim', '')[:60]}...")



def export_to_excel(output_jsonl: str, excel_path: str) -> None:
    """Read output.jsonl and write long_answer + revised_claim to Excel."""
    try:
        results = []
        with jsonlines.open(output_jsonl) as reader:
            for obj in reader:
                results.append(obj)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REVISE Results"

        # Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", start_color="1F4E79")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Headers
        headers = ["#", "Question", "Long Answer (Ground Truth)", "Revised Claim"]
        col_widths = [5, 30, 60, 60]
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 25

        # Alternating row colors
        fill_light = PatternFill("solid", start_color="EBF3FB")
        fill_white = PatternFill("solid", start_color="FFFFFF")

        for i, obj in enumerate(results, start=1):
            row = i + 1
            fill = fill_light if i % 2 == 0 else fill_white
            values = [
                i,
                obj.get("question_text", ""),
                obj.get("long_answer", ""),
                obj.get("revised_claim", ""),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = center_align if col == 1 else wrap_align
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Arial", size=10)

            ws.row_dimensions[row].height = 80

        wb.save(excel_path)
        logger.info(f"Excel saved to {excel_path}")

    except Exception as e:
        logger.error(f"Failed to create Excel file: {e}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    args = get_args()
    logger.info(f"Reading claims from: {args.input_file}")
    claims = read_claims(args.input_file)
    logger.info(f"Total claims loaded: {len(claims)}")

    # Limit to first 10 for testing
    claims = claims[:10]
    logger.info(f"Processing first {len(claims)} claims for test run.")

    revise_and_save_jsonl(
        claims=claims,
        output_jsonl=args.output_file,
        model=args.model,
        temperature_qgen=args.temperature_qgen,
        max_search_results_per_query=args.max_search_results_per_query,
        max_sentences_per_passage=args.max_sentences_per_passage,
        sliding_distance=args.sliding_distance,
        max_passages_per_search_result=args.max_passages_per_search_result,
        hallucinate_evidence=args.hallucinate_evidence,
    )
    logger.info(f"Done. Output saved to {args.output_file}")

    # Export to Excel
    excel_path = args.output_file.replace(".jsonl", ".xlsx")
    logger.info(f"Exporting results to Excel: {excel_path}")
    export_to_excel(args.output_file, excel_path)


if __name__ == "__main__":
    main()
# import argparse
# import json
# import logging
# import re
# from typing import Any, Dict, List, Optional
# import warnings
# warnings.filterwarnings("ignore", message="Can't initialize NVML")

# import csv
# import jsonlines
# import Levenshtein
# import nltk
# from tqdm import tqdm
# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# nltk.download('punkt')

# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler("run_editor_sequential.log"),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)

# def get_args() -> argparse.Namespace:
#     """Parses command line arguments."""
#     parser = argparse.ArgumentParser(description="Revise claims from a JSONL file and output to JSONL.")
#     parser.add_argument("--input_file", type=str, required=True, help="JSONLines file of claims.")
#     parser.add_argument("--output_file", type=str, required=True, help="Output JSONL file.")
#     parser.add_argument("--model", default="gpt-3.5-turbo", type=str, help="GPT model to use.")
#     parser.add_argument("--temperature_qgen", default=0.7, type=float, help="Sampling temperature for question generation.")
#     parser.add_argument("--hallucinate_evidence", action="store_true", help="Use hallucinated evidence instead of retrieval.")
#     parser.add_argument("--max_search_results_per_query", default=5, type=int, help="Max search results per query.")
#     parser.add_argument("--max_sentences_per_passage", default=4, type=int, help="Max sentences per evidence passage.")
#     parser.add_argument("--sliding_distance", default=1, type=int, help="Sliding window distance for passage extraction.")
#     parser.add_argument("--max_passages_per_search_result", default=1, type=int, help="Max passages per search result.")
#     return parser.parse_args()

# def split_into_atomic_statements(text: str) -> List[str]:
#     """Splits text into atomic statements using NLTK's sentence tokenizer."""
#     text = text.strip()
#     return nltk.sent_tokenize(text) if text else []

# def process_atomic_statement(
#     statement: str,
#     accumulated_context: str,
#     model: str,
#     temperature_qgen: float,
#     search_params: Dict[str, Any],
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     """
#     Process a single atomic statement: question generation, evidence retrieval, and revision.
#     """
#     try:
#         questions = question_generation.run_rarr_question_generation(
#             claim=statement,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if accumulated_context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             context=accumulated_context,
#             required_questions=3
#         )
        
#         all_evidences = []
#         revised_statement = statement
        
#         for question in questions:
#             evidences = search.run_search(
#                 query=question,
#                 context=accumulated_context,
#                 atomic_statement=statement,
#                 **search_params
#             )
            
#             for evidence_item in evidences:
#                 gate_result = agreement_gate.run_agreement_gate(
#                     claim=revised_statement,
#                     query=question,
#                     evidence=evidence_item["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if accumulated_context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                     context=accumulated_context
#                 )
#                 all_evidences.append({"text": evidence_item["text"], "query": question, "gate_result": gate_result})
                
#                 if gate_result.get("is_open"):
#                     edited_result = editor.run_rarr_editor(
#                         claim=revised_statement,
#                         query=question,
#                         evidence=evidence_item["text"],
#                         model=model,
#                         prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if accumulated_context else rarr_prompts.EDITOR_PROMPT,
#                         context=accumulated_context
#                     )
#                     if edited_result.get("text"):
#                         revised_statement = edited_result["text"]
        
#         return {
#             "original_statement": statement,
#             "revised_statement": revised_statement,
#             "questions": questions,
#             "evidences": all_evidences
#         }
#     except Exception as e:
#         logger.error(f"Error processing statement: {e}")
#         return {"original_statement": statement, "revised_statement": statement, "questions": [], "evidences": []}

# def main():
#     args = get_args()
#     logger.info(f"Reading claims from JSONL file: {args.input_file}")
#     claims = read_claims(args.input_file)
#     logger.info(f"Total claims read: {len(claims)}")

#     logger.info("Revising claims and writing to JSONL...")
#     revise_and_save_jsonl(
#         claims=claims,
#         output_jsonl=args.output_file,
#         model=args.model,
#         temperature_qgen=args.temperature_qgen,
#         max_search_results_per_query=args.max_search_results_per_query,
#         max_sentences_per_passage=args.max_sentences_per_passage,
#         sliding_distance=args.sliding_distance,
#         max_passages_per_search_result=args.max_passages_per_search_result,
#         hallucinate_evidence=args.hallucinate_evidence
#     )

#     logger.info(f"Done. Output saved to {args.output_file}.")

# if __name__ == "__main__":
#     main()


# import torch
# from transformers import AutoModelForSequenceClassification, AutoTokenizer
# import Levenshtein
# import numpy as np
# import argparse
# import json
# import os
# import warnings
# from typing import Any, Dict
# from datetime import datetime

# import jsonlines
# import tqdm
# import logging
# import openai
# from time import time

# # Suppress the specific FutureWarning from transformers
# warnings.filterwarnings("ignore", category=FutureWarning, module="transformers")

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# # Import necessary functions and prompts
# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# # Load the NLI model and tokenizer
# nli_model_name = "roberta-large-mnli"
# nli_model = AutoModelForSequenceClassification.from_pretrained(nli_model_name)
# tokenizer = AutoTokenizer.from_pretrained(nli_model_name)

# def setup_openai_client():
#     """Set up the OpenAI client."""
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def calculate_evidence_reliability(evidence: Dict[str, Any]) -> float:
#     """Calculate a reliability score for the evidence."""
#     current_year = datetime.now().year
#     age_penalty = max(0, current_year - int(evidence.get('year', current_year))) * 0.1
#     return min(1.0, evidence.get('retrieval_score', 0) / 10 - age_penalty)

# def calculate_preslev(original_text: str, revised_text: str) -> float:
#     """Calculates the PresLev metric for preservation based on Levenshtein distance."""
#     if not original_text:  # Handle empty original text case
#         return 0.0

#     lev_distance = Levenshtein.distance(original_text, revised_text)
#     preslev_score = max(1 - lev_distance / len(original_text), 0)  # Score remains between 0 and 1
#     return preslev_score

# def calculate_attrauto(revised_text: str, evidences: list) -> float:
#     """Enhanced Attrauto calculation"""
#     sentences = [s.strip() for s in revised_text.split('.') if s.strip()]
#     entailment_scores = []

#     for sentence in sentences:
#         max_entailment_score = 0.0
        
#         # Combine all evidence texts for better context
#         combined_evidence = " ".join([e['text'] for e in evidences])
        
#         # Calculate entailment with combined evidence
#         inputs = tokenizer.encode_plus(
#             sentence, 
#             combined_evidence, 
#             return_tensors="pt",
#             truncation=True,
#             max_length=512
#         )
        
#         with torch.no_grad():
#             outputs = nli_model(**inputs)
#             probs = torch.softmax(outputs.logits[0], dim=0)
#             entailment_prob = probs[2].item()  # Entailment class
#             max_entailment_score = max(max_entailment_score, entailment_prob)
            
#         entailment_scores.append(max_entailment_score)

#     return np.mean(entailment_scores) if entailment_scores else 0.0


# # def calculate_attrauto(revised_text: str, evidences: list) -> float:
# #     """
# #     Calculates the Attrauto metric for attribution using an NLI model.
    
# #     Args:
# #         revised_text (str): The text to attribute (e.g., claim or generated text).
# #         evidences (list of str): List of evidence snippets.

# #     Returns:
# #         float: The Attrauto score.
# #     """
# #     sentences = revised_text.split('.')
# #     entailment_scores = []

# #     for sentence in sentences:
# #         sentence = sentence.strip()
# #         if not sentence:
# #             continue

# #         max_entailment_score = 0.0  # Start with 0 as the minimum possible score

# #         for evidence in evidences:
# #             # Decontextualize the sentence if needed (to be implemented as per your requirements)
# #             # Here, we assume the sentence is already decontextualized.
            
# #             # Calculate NLI entailment probability for this evidence-sentence pair
# #             inputs = tokenizer.encode_plus(sentence, evidence['text'], return_tensors="pt", truncation=True)
# #             with torch.no_grad():
# #                 outputs = nli_model(**inputs)
# #             entailment_logits = outputs.logits[0]
# #             entailment_prob = torch.softmax(entailment_logits, dim=0)[2].item()  # Index 2 is for "entailment" class

# #             # Update max entailment score for this sentence
# #             max_entailment_score = max(max_entailment_score, entailment_prob)

# #         entailment_scores.append(max_entailment_score)

# #     # Calculate the average of the maximum entailment scores per sentence for the final Attrauto score
# #     attrauto_score = np.mean(entailment_scores) if entailment_scores else 0.0
# #     return attrauto_score

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.8,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 6,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 50,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     """Run the RARR editor on a single instance (claim)."""
#     try:
#         start_time = time()  # Track computation time
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Step 1: Generate questions for the claim
#         questions = question_generation.run_rarr_question_generation(
#             claim=claim,
#             context=context,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             num_rounds=num_rounds_qgen,
#             client=client,
#         )

#         # Step 2: Handle cases where questions generated are fewer than 5
#         if not questions or len(questions) < 5:
#             logging.warning(f"Generated only {len(questions)} questions. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": questions,
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Step 3: Run search for each generated question and retrieve evidence
#         evidences_for_questions = []
#         for query in questions:
#             evidences = search.run_search(
#                 query=query,
#                 max_search_results_per_query=max_search_results_per_query,
#                 max_sentences_per_passage=max_sentences_per_passage,
#                 sliding_distance=sliding_distance,
#                 max_passages_per_search_result_to_return=max_passages_per_search_result,
#             )
#             evidences_for_questions.append(evidences)

#         # Step 4: Flatten the evidences and calculate reliability scores
#         used_evidences = []
#         for cur_evids in evidences_for_questions:
#             for e in cur_evids[:max_evidences_per_question]:
#                 e['reliability_score'] = calculate_evidence_reliability(e)
#                 used_evidences.append(e)

#         # Step 5: Iterative editing over each evidence
#         revisions = []
#         for evid in used_evidences:
#             # Run the agreement gate for each question-evidence pair
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # If the agreement gate is open, run the editor
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Check if the edit ratio is within the allowed threshold
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     edit_distance = Levenshtein.distance(original_claim, edited_claim)
                    
#                     # Log the Levenshtein distance
#                     logging.info(f"Levenshtein distance between original claim and edited claim: {edit_distance}")
                    
#                     revisions.append({
#                         "text": edited_claim,
#                         "evidence_reliability": evid.get('reliability_score', 0),
#                         "edit_distance": edit_distance
#                     })

#             # Add the claim to revision steps
#             revision_steps.append({"text": claim})

#         # Step 6: Select the best revision based on reliability and edit distance
#         if revisions:
#             best_revision = max(revisions, key=lambda r: r['evidence_reliability'] - r['edit_distance'] / len(original_claim))
#             final_text = best_revision['text']
#         else:
#             final_text = original_claim

#         # Step 7: Calculate Attrauto and PresLev scores
#         preslev_score = calculate_preslev(original_claim, final_text)
#         attrauto_score = calculate_attrauto(final_text, used_evidences)
        
#         logging.info(f"PresLev score: {preslev_score:.4f}")
#         logging.info(f"Attrauto score: {attrauto_score:.4f}")

#         # Step 8: Prepare the result object with selected evidences
#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": final_text,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                     "PresLev": preslev_score,
#                     "Attrauto": attrauto_score
#                 }
#             ],
#         }

#         # Select up to 5 evidences for the final result
#         selected_evidences = evidence_selection.select_evidences(result, max_selected=5)
#         result["selected_evidences"] = selected_evidences

#         end_time = time()  # Log computation time
#         logging.info(f"Instance processing time: {end_time - start_time:.2f} seconds")

#         return result

#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.9,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If set, hallucinate evidence instead of retrieving it (for experiments).",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results to retrieve per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=5,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=5,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     return parser.parse_args()

# def main() -> None:
#     """Main function to load claims and run the RARR editor."""
#     args = get_args()

#     # Load finished results if resuming
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()




































































































































# import argparse
# import json
# import os
# from typing import Any, Dict
# from datetime import datetime

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# # Import necessary functions and prompts
# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     """Set up the OpenAI client."""
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     """Warn the user if hallucinating evidence."""
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def calculate_evidence_reliability(evidence: Dict[str, Any]) -> float:
#     """Calculate a reliability score for the evidence."""
#     current_year = datetime.now().year
#     age_penalty = max(0, current_year - int(evidence.get('year', current_year))) * 0.1
#     return min(1.0, evidence.get('retrieval_score', 0) / 10 - age_penalty)

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     """Run the RARR editor on a single instance (claim)."""
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Step 1: Generate questions for the claim
#         questions = question_generation.run_rarr_question_generation(
#             claim=claim,
#             context=context,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             num_rounds=num_rounds_qgen,
#             client=client,
#         )

#         # Step 2: Handle cases where questions generated are fewer than 5
#         if not questions or len(questions) < 5:
#             logging.warning(f"Generated only {len(questions)} questions. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": questions,
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Step 3: Run search for each generated question and retrieve evidence
#         evidences_for_questions = []
#         for query in questions:
#             evidences = search.run_search(
#                 query=query,
#                 max_search_results_per_query=max_search_results_per_query,
#                 max_sentences_per_passage=max_sentences_per_passage,
#                 sliding_distance=sliding_distance,
#                 max_passages_per_search_result_to_return=max_passages_per_search_result,
#             )
#             evidences_for_questions.append(evidences)

#         # Step 4: For each question, calculate the min and max retrieval scores
#         for idx, evidences in enumerate(evidences_for_questions):
#             # Extract the retrieval scores for the current question-evidence pair
#             scores = [evidence['retrieval_score'] for evidence in evidences]

#             # Calculate the minimum and maximum retrieval scores
#             if scores:  # Ensure scores exist
#                 min_score = min(scores)
#                 max_score = max(scores)
#             else:
#                 min_score = None
#                 max_score = None

#             # Log the retrieval score range for each question
#             logging.info(f"Question {idx + 1}: '{questions[idx]}'")
#             logging.info(f"Min retrieval score: {min_score}")
#             logging.info(f"Max retrieval score: {max_score}")

#         # Step 5: Flatten the evidences and calculate reliability scores
#         used_evidences = []
#         for cur_evids in evidences_for_questions:
#             for e in cur_evids[:max_evidences_per_question]:
#                 e['reliability_score'] = calculate_evidence_reliability(e)
#                 used_evidences.append(e)

#         # Step 6: Sort evidences by reliability score
#         used_evidences.sort(key=lambda x: x['reliability_score'], reverse=True)

#         # Step 7: Iterative editing over each evidence
#         revisions = []
#         for evid in used_evidences:
#             # Run the agreement gate for each question-evidence pair
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # If the agreement gate is open, run the editor
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Check if the edit ratio is within the allowed threshold
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     edit_distance = Levenshtein.distance(original_claim, edited_claim)
                    
#                     # Print or log the Levenshtein distance
#                     print(f"Levenshtein distance between original claim and edited claim: {edit_distance}")
#                     logging.info(f"Levenshtein distance between original claim and edited claim: {edit_distance}")
                    
#                     revisions.append({
#                         "text": edited_claim,
#                         "evidence_reliability": evid.get('reliability_score', 0),
#                         "edit_distance": edit_distance
#                     })


#             # Add the claim to revision steps
#             revision_steps.append({"text": claim})

#         # Step 8: Select the best revision based on reliability and edit distance
#         if revisions:
#             best_revision = max(revisions, key=lambda r: r['evidence_reliability'] - r['edit_distance'] / len(original_claim))
#             final_text = best_revision['text']
#         else:
#             final_text = original_claim

#         # Step 9: Prepare the result object with selected evidences
#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": final_text,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ],
#         }

#         # Select up to 5 evidences for the final result
#         selected_evidences = evidence_selection.select_evidences(result, max_selected=5)
#         result["selected_evidences"] = selected_evidences

#         return result

#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }


# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If set, hallucinate evidence instead of retrieving it (for experiments).",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results to retrieve per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     return parser.parse_args()

# def main() -> None:
#     """Main function to load claims and run the RARR editor."""
#     args = get_args()

#     # Load finished results if resuming
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()

# import argparse
# import json
# import os
# from typing import Any, Dict, List
# from datetime import datetime

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# # Import necessary functions and prompts
# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     """Set up the OpenAI client."""
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     """Warn the user if hallucinating evidence."""
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def calculate_evidence_reliability(evidence: Dict[str, Any]) -> float:
#     """Calculate a reliability score for the evidence."""
#     current_year = datetime.now().year
#     age_penalty = max(0, current_year - int(evidence.get('year', current_year))) * 0.1
#     return min(1.0, evidence.get('retrieval_score', 0) / 10 - age_penalty)

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     """Run the RARR editor on a single instance (claim)."""
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Generate questions for the claim
#         questions = question_generation.run_rarr_question_generation(
#             claim=claim,
#             context=context,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             num_rounds=num_rounds_qgen,
#             client=client,
#         )

#         if not questions or len(questions) < 5:
#             logging.warning(f"Generated only {len(questions)} questions. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": questions,
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Run search or hallucinate evidence
#         if hallucinate_evidence:
#             raise_hallucinate_evidence_warning()
#             evidences_for_questions = [
#                 [
#                     hallucination.run_evidence_hallucination(
#                         query=query,
#                         model=model,
#                         prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                         client=client,
#                     )
#                 ]
#                 for query in questions
#             ]
#         else:
#             evidences_for_questions = [
#                 search.run_search(
#                     query=query,
#                     max_search_results_per_query=max_search_results_per_query,
#                     max_sentences_per_passage=max_sentences_per_passage,
#                     sliding_distance=sliding_distance,
#                     max_passages_per_search_result_to_return=max_passages_per_search_result,
#                 )
#                 for query in questions
#             ]

#         # Flatten the evidences and calculate reliability scores
#         used_evidences = []
#         for cur_evids in evidences_for_questions:
#             for e in cur_evids[:max_evidences_per_question]:
#                 e['reliability_score'] = calculate_evidence_reliability(e)
#                 used_evidences.append(e)

#         # Sort evidences by reliability score
#         used_evidences.sort(key=lambda x: x['reliability_score'], reverse=True)

#         # Iterative editing over each evidence
#         revisions = []
#         for evid in used_evidences:
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # Run the editor if the agreement gate is open
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Check if the edit is within the allowed ratio
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     revisions.append({
#                         "text": edited_claim,
#                         "evidence_reliability": evid.get('reliability_score', 0),
#                         "edit_distance": Levenshtein.distance(original_claim, edited_claim)
#                     })

#             revision_steps.append({"text": claim})

#         # Select the best revision
#         if revisions:
#             best_revision = max(revisions, key=lambda r: r['evidence_reliability'] - r['edit_distance'] / len(original_claim))
#             final_text = best_revision['text']
#         else:
#             final_text = original_claim

#         # Prepare the result
#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": final_text,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ],
#         }

#         # Select up to 5 evidences
#         selected_evidences = evidence_selection.select_evidences(result, max_selected=5)
#         result["selected_evidences"] = selected_evidences

#         return result
#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If set, hallucinate evidence instead of retrieving it (for experiments).",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results to retrieve per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     return parser.parse_args()

# def main() -> None:
#     """Main function to load claims and run the RARR editor."""
#     args = get_args()

#     # Load finished results if resuming
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()




# """Runs the RARR editor on a JSONL file of claims.

# Runs question generation, retrieval, agreement gate, and editing on a file with claims
# using GPT-3 and Bing.
# """
# import argparse
# import json
# import os
# from typing import Any, Dict, List
# from datetime import datetime

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def calculate_evidence_reliability(evidence: Dict[str, Any]) -> float:
#     # This is a placeholder function. In a real scenario, you'd implement
#     # a more sophisticated reliability scoring system.
#     current_year = datetime.now().year
#     age_penalty = max(0, current_year - int(evidence.get('year', current_year))) * 0.1
#     return min(1.0, evidence.get('retrieval_score', 0) / 10 - age_penalty)

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Generate questions for the claim
#         questions = question_generation.run_rarr_question_generation(
#             claim=claim,
#             context=context,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             num_rounds=num_rounds_qgen,
#             client=client,
#         )

#         if not questions or len(questions) < 5:
#             logging.warning(f"Generated only {len(questions)} questions. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": questions,
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Run search on generated question for the claim
#         if hallucinate_evidence:
#             raise_hallucinate_evidence_warning()
#             evidences_for_questions = [
#                 [
#                     hallucination.run_evidence_hallucination(
#                         query=query,
#                         model=model,
#                         prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                         client=client,
#                     )
#                 ]
#                 for query in questions
#             ]
#         else:
#             evidences_for_questions = [
#                 search.run_search(
#                     query=query,
#                     max_search_results_per_query=max_search_results_per_query,
#                     max_sentences_per_passage=max_sentences_per_passage,
#                     sliding_distance=sliding_distance,
#                     max_passages_per_search_result_to_return=max_passages_per_search_result,
#                 )
#                 for query in questions
#             ]

#         # Flatten the evidences per question into a single list and calculate reliability
#         used_evidences = []
#         for cur_evids in evidences_for_questions:
#             for e in cur_evids[:max_evidences_per_question]:
#                 e['reliability_score'] = calculate_evidence_reliability(e)
#                 used_evidences.append(e)

#         # Sort evidences by reliability score
#         used_evidences.sort(key=lambda x: x['reliability_score'], reverse=True)

#         # Iterative editing over each evidence
#         revisions = []
#         for evid in used_evidences:
#             # Run the agreement gate on the current (claim, context, query, evidence) tuple
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # Run the editor gate if the agreement gate is open
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Don't keep the edit if the editor makes a huge change
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     revisions.append({
#                         "text": edited_claim,
#                         "evidence_reliability": evid.get('reliability_score', 0),
#                         "edit_distance": Levenshtein.distance(original_claim, edited_claim)
#                     })

#             revision_steps.append({"text": claim})

#         # Select the best revision
#         if revisions:
#             best_revision = max(revisions, key=lambda r: r['evidence_reliability'] - r['edit_distance']/len(original_claim))
#             final_text = best_revision['text']
#         else:
#             final_text = original_claim

#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": final_text,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ],
#         }

#         # Ensure we select 5 evidences (or max available)
#         selected_evidences = evidence_selection.select_evidences(result, max_selected=5)
#         result["selected_evidences"] = selected_evidences

#         return result
#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If this flag is set, we hallucinate evidence instead of retrieving it. "
#         "This flag should NEVER be set when trying to improve attribution as evidence  "
#         "may be inaccurate and is only provided to quickly experiment with repository "
#         "setting up the search API first.",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results we get per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages from a search result.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result. A passage"
#         " ranker is applied to get the top passages per query.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     args = parser.parse_args()

#     # Write all args to file
#     with open(args.output_file + "_args", "w", encoding="utf-8") as writer:
#         json.dump(args.__dict__, writer, indent=4)
#     return args

# def main() -> None:
#     """Loads a RARR evaluation set and runs GPT-3 RARR editing."""
#     args = get_args()

#     # Load the finished results by mapping from the claim name to the results.
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()






# """Runs the RARR editor on a JSONL file of claims.

# Runs question generation, retrieval, agreement gate, and editing on a file with claims
# using GPT-3 and Bing.
# """
# import argparse
# import json
# import os
# from typing import Any, Dict, List

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Generate questions for the claim
#         unique_questions = set()
#         max_attempts = num_rounds_qgen * 2  # Allow double the attempts to get unique questions
#         attempts = 0

#         while len(unique_questions) < num_rounds_qgen and attempts < max_attempts:
#             new_questions = question_generation.run_rarr_question_generation(
#                 claim=claim,
#                 context=context,
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#                 temperature=temperature_qgen,
#                 num_rounds=1,  # Generate one round at a time
#                 client=client,
#             )
#             unique_questions.update(new_questions)  # Add new questions to the set
#             attempts += 1

#         questions = list(unique_questions)[:num_rounds_qgen]  # Take the first 5 unique questions

#         if not questions:
#             logging.warning("No questions generated. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": [],
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Run search on generated question for the claim
#         if hallucinate_evidence:
#             raise_hallucinate_evidence_warning()
#             evidences_for_questions = [
#                 [
#                     hallucination.run_evidence_hallucination(
#                         query=query,
#                         model=model,
#                         prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                         client=client,
#                     )
#                 ]
#                 for query in questions
#             ]
#         else:
#             evidences_for_questions = [
#                 search.run_search(
#                     query=query,
#                     max_search_results_per_query=max_search_results_per_query,
#                     max_sentences_per_passage=max_sentences_per_passage,
#                     sliding_distance=sliding_distance,
#                     max_passages_per_search_result_to_return=max_passages_per_search_result,
#                 )
#                 for query in questions
#             ]

#         # Flatten the evidences per question into a single list.
#         used_evidences = [
#             e
#             for cur_evids in evidences_for_questions
#             for e in cur_evids[:max_evidences_per_question]
#         ]

#         # Iterative editing over each evidence
#         for evid in used_evidences:
#             # Run the agreement gate on the current (claim, context, query, evidence) tuple
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # Run the editor gate if the agreement gate is open
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Don't keep the edit if the editor makes a huge change
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     claim = edited_claim

#             revision_steps.append({"text": claim})

#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": revision_steps[-1]["text"] if revision_steps else original_claim,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ] if revision_steps else [],
#         }
#         selected_evidences = evidence_selection.select_evidences(result)
#         result["selected_evidences"] = selected_evidences
#         return result
#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,  # Changed from 3 to 5
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If this flag is set, we hallucinate evidence instead of retrieving it. "
#         "This flag should NEVER be set when trying to improve attribution as evidence  "
#         "may be inaccurate and is only provided to quickly experiment with repository "
#         "setting up the search API first.",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results we get per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages from a search result.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result. A passage"
#         " ranker is applied to get the top passages per query.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     args = parser.parse_args()

#     # Write all args to file
#     with open(args.output_file + "_args", "w", encoding="utf-8") as writer:
#         json.dump(args.__dict__, writer, indent=4)
#     return args

# def main() -> None:
#     """Loads a RARR evaluation set and runs GPT-3 RARR editing."""
#     args = get_args()

#     # Load the finished results by mapping from the claim name to the results.
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()
# """Runs the RARR editor on a JSONL file of claims.

# Runs question generation, retrieval, agreement gate, and editing on a file with claims
# using GPT-3 and Bing.
# """
# import argparse
# import json
# import os
# from typing import Any, Dict, List

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Generate questions for the claim
#         unique_questions = set()
#         max_attempts = num_rounds_qgen * 2  # Allow double the attempts to get unique questions
#         attempts = 0

#         while len(unique_questions) < num_rounds_qgen and attempts < max_attempts:
#             new_questions = question_generation.run_rarr_question_generation(
#                 claim=claim,
#                 context=context,
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#                 temperature=temperature_qgen,
#                 num_rounds=1,  # Generate one round at a time
#                 client=client,
#             )
#             unique_questions.update(new_questions)  # Add new questions to the set
#             attempts += 1

#         questions = list(unique_questions)[:num_rounds_qgen]  # Take the first 5 unique questions

#         if not questions:
#             logging.warning("No questions generated. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": [],
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Run search on generated question for the claim
#         if hallucinate_evidence:
#             raise_hallucinate_evidence_warning()
#             evidences_for_questions = [
#                 [
#                     hallucination.run_evidence_hallucination(
#                         query=query,
#                         model=model,
#                         prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                         client=client,
#                     )
#                 ]
#                 for query in questions
#             ]
#         else:
#             evidences_for_questions = [
#                 search.run_search(
#                     query=query,
#                     max_search_results_per_query=max_search_results_per_query,
#                     max_sentences_per_passage=max_sentences_per_passage,
#                     sliding_distance=sliding_distance,
#                     max_passages_per_search_result_to_return=max_passages_per_search_result,
#                 )
#                 for query in questions
#             ]

#         # Flatten the evidences per question into a single list.
#         used_evidences = [
#             e
#             for cur_evids in evidences_for_questions
#             for e in cur_evids[:max_evidences_per_question]
#         ]

#         # Iterative editing over each evidence
#         for evid in used_evidences:
#             # Run the agreement gate on the current (claim, context, query, evidence) tuple
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # Run the editor gate if the agreement gate is open
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Don't keep the edit if the editor makes a huge change
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     claim = edited_claim

#             revision_steps.append({"text": claim})

#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": revision_steps[-1]["text"] if revision_steps else original_claim,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ] if revision_steps else [],
#         }
#         selected_evidences = evidence_selection.select_evidences(result)
#         result["selected_evidences"] = selected_evidences
#         return result
#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,  # Changed from 3 to 5
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If this flag is set, we hallucinate evidence instead of retrieving it. "
#         "This flag should NEVER be set when trying to improve attribution as evidence  "
#         "may be inaccurate and is only provided to quickly experiment with repository "
#         "setting up the search API first.",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results we get per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages from a search result.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result. A passage"
#         " ranker is applied to get the top passages per query.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     args = parser.parse_args()

#     # Write all args to file
#     with open(args.output_file + "_args", "w", encoding="utf-8") as writer:
#         json.dump(args.__dict__, writer, indent=4)
#     return args

# def main() -> None:
#     """Loads a RARR evaluation set and runs GPT-3 RARR editing."""
#     args = get_args()

#     # Load the finished results by mapping from the claim name to the results.
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()
# """Runs the RARR editor on a JSONL file of claims.

# Runs question generation, retrieval, agreement gate, and editing on a file with claims
# using GPT-3 and Bing.
# """
# import argparse
# import json
# import os
# from typing import Any, Dict, List

# import jsonlines
# import Levenshtein
# import tqdm
# import logging
# import openai

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )

# def setup_openai_client():
#     api_key = os.getenv("OPENAI_API_KEY")
#     if not api_key:
#         raise ValueError("OPENAI_API_KEY environment variable is not set")

#     if hasattr(openai, 'OpenAI'):
#         # New version of the library
#         return openai.OpenAI(api_key=api_key)
#     else:
#         # Old version of the library
#         openai.api_key = api_key
#         return openai

# # Initialize the OpenAI client
# client = setup_openai_client()

# def raise_hallucinate_evidence_warning():
#     if not hasattr(raise_hallucinate_evidence_warning, "called"):
#         logging.warning(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#         raise_hallucinate_evidence_warning.called = True

# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = "gpt-3.5-turbo",
#     temperature_qgen: float = 0.9,
#     num_rounds_qgen: int = 5,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     try:
#         original_claim = claim
#         agreement_gates = []
#         revision_steps = []

#         # Generate questions for the claim
#         questions = question_generation.run_rarr_question_generation(
#             claim=claim,
#             context=context,
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT if context else rarr_prompts.QGEN_PROMPT,
#             temperature=temperature_qgen,
#             num_rounds=num_rounds_qgen,
#             client=client,
#         )

#         if not questions:
#             logging.warning("No questions generated. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": [],
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#         # Run search on generated question for the claim
#         if hallucinate_evidence:
#             raise_hallucinate_evidence_warning()
#             evidences_for_questions = [
#                 [
#                     hallucination.run_evidence_hallucination(
#                         query=query,
#                         model=model,
#                         prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                         client=client,
#                     )
#                 ]
#                 for query in questions
#             ]
#         else:
#             evidences_for_questions = [
#                 search.run_search(
#                     query=query,
#                     max_search_results_per_query=max_search_results_per_query,
#                     max_sentences_per_passage=max_sentences_per_passage,
#                     sliding_distance=sliding_distance,
#                     max_passages_per_search_result_to_return=max_passages_per_search_result,
#                 )
#                 for query in questions
#             ]

#         # Flatten the evidences per question into a single list.
#         used_evidences = [
#             e
#             for cur_evids in evidences_for_questions
#             for e in cur_evids[:max_evidences_per_question]
#         ]

#         # Iterative editing over each evidence
#         for evid in used_evidences:
#             # Run the agreement gate on the current (claim, context, query, evidence) tuple
#             gate = agreement_gate.run_agreement_gate(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT if context else rarr_prompts.AGREEMENT_GATE_PROMPT,
#                 client=client,
#             )
#             agreement_gates.append(gate)

#             # Run the editor gate if the agreement gate is open
#             if gate["is_open"]:
#                 edited_claim = editor.run_rarr_editor(
#                     claim=claim,
#                     context=context,
#                     query=evid["query"],
#                     evidence=evid["text"],
#                     model=model,
#                     prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT if context else rarr_prompts.EDITOR_PROMPT,
#                     client=client,
#                 )["text"]

#                 # Don't keep the edit if the editor makes a huge change
#                 if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                     claim = edited_claim

#             revision_steps.append({"text": claim})

#         result = {
#             "context": context,
#             "text": original_claim,
#             "questions": questions,
#             "evidences_for_questions": evidences_for_questions,
#             "revisions": [
#                 {
#                     "original_text": original_claim,
#                     "revised_text": revision_steps[-1]["text"] if revision_steps else original_claim,
#                     "evidences": used_evidences,
#                     "agreement_gates": agreement_gates,
#                     "revision_steps": revision_steps,
#                 }
#             ] if revision_steps else [],
#         }
#         selected_evidences = evidence_selection.select_evidences(result)
#         result["selected_evidences"] = selected_evidences
#         return result
#     except Exception as e:
#         logging.error(f"Error in run_editor_one_instance: {str(e)}")
#         return {
#             "context": context,
#             "text": claim,
#             "error": str(e)
#         }

# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default="gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=5,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If this flag is set, we hallucinate evidence instead of retrieving it. "
#         "This flag should NEVER be set when trying to improve attribution as evidence  "
#         "may be inaccurate and is only provided to quickly experiment with repository "
#         "setting up the search API first.",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results we get per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages from a search result.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result. A passage"
#         " ranker is applied to get the top passages per query.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     args = parser.parse_args()

#     # Write all args to file
#     with open(args.output_file + "_args", "w", encoding="utf-8") as writer:
#         json.dump(args.__dict__, writer, indent=4)
#     return args

# def main() -> None:
#     """Loads a RARR evaluation set and runs GPT-3 RARR editing."""
#     args = get_args()

#     # Load the finished results by mapping from the claim name to the results.
#     if args.resume and os.path.exists(args.output_file):
#         logging.info(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         logging.info(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             try:
#                 claim = line["input_info"][args.claim_field]
#                 context = None
#                 if args.context_field:
#                     context = line["input_info"].get(args.context_field, "")
#                     context = " ".join(context.split("\n")) if context else None

#                 # Search for finished result
#                 if finished_results and claim in finished_results:
#                     line["result"] = finished_results[claim]
#                 else:
#                     line["result"] = run_editor_one_instance(
#                         model=args.model,
#                         claim=claim,
#                         context=context,
#                         temperature_qgen=args.temperature_qgen,
#                         num_rounds_qgen=args.num_rounds_qgen,
#                         max_search_results_per_query=args.max_search_results_per_query,
#                         max_sentences_per_passage=args.max_sentences_per_passage,
#                         sliding_distance=args.sliding_distance,
#                         max_passages_per_search_result=args.max_passages_per_search_result,
#                         max_evidences_per_question=args.max_evidences_per_question,
#                         max_edit_ratio=args.max_edit_ratio,
#                         hallucinate_evidence=args.hallucinate_evidence,
#                     )
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")
#             except Exception as e:
#                 logging.error(f"Error processing line: {str(e)}")
#                 line["error"] = str(e)
#                 writer.write(json.dumps(line, ensure_ascii=False) + "\n")

# if __name__ == "__main__":
#     main()
# """Runs the RARR editor on a JSONL file of claims.

# Runs question generation, retrieval, agreement gate, and editing on a file with claims
# using GPT-3 and Bing.
# """
# import argparse
# import json
# import os
# from typing import Any, Dict

# import jsonlines
# import Levenshtein
# import tqdm
# import logging

# logging.basicConfig(level=logging.INFO)

# from prompts import hallucination_prompts, rarr_prompts
# from utils import (
#     agreement_gate,
#     editor,
#     evidence_selection,
#     hallucination,
#     search,
#     question_generation,
# )


# def raise_hallucinate_evidence_warning():
#     if not raise_hallucinate_evidence_warning.called:
#         print(
#             "WARNING!! YOU ARE USING A LLM TO GENERATE EVIDENCE POTENTIALLY WITH "
#             "HALLUCINATIONS INSTEAD OF RETRIEVING EVIDENCE. \n\nThis should NEVER be "
#             "done when trying to improve attribution as evidence may be inaccurate "
#             "and is only provided to quickly experiment with repository setting up "
#             "the search API first.\n"
#         )
#     raise_hallucinate_evidence_warning.called = True


# raise_hallucinate_evidence_warning.called = False


# def run_editor_one_instance(
#     claim: str,
#     context: str = None,
#     model: str = " gpt-3.5-turbo",
#     temperature_qgen: float = 0.7,
#     num_rounds_qgen: int = 3,
#     max_search_results_per_query: int = 5,
#     max_sentences_per_passage: int = 4,
#     sliding_distance: int = 1,
#     max_passages_per_search_result: int = 1,
#     max_evidences_per_question: int = 1,
#     max_edit_ratio: float = 100,
#     hallucinate_evidence: bool = False,
# ) -> Dict[str, Any]:
#     """Runs query generation, search, agreement gating, and editing on a claim.

#     Args:
#         claim: Text to check the validity of.
#         model: Name of the OpenAI GPT-3 model to use.
#         temperature_qgen: Sampling temperature to use for query generation.
#         num_rounds_qgen: Number of times to sample questions.
#         max_search_results_per_query: Maximum number of search results per query.
#         max_sentences_per_passage: Maximum number of sentences for each passage.
#         sliding_distance: Sliding window distance over the sentences of each search
#             result. Used to extract passages.
#         max_passages_per_search_result:  Maximum number of passages to return for
#             each search result. A passage ranker is applied first.
#         max_evidences_per_question: Maximum number of evidences to return per question.
#         max_edit_ratio: Maximum edit ratio between claim and edit for each round.
#     Returns:
#         result: All revision information, including the queries generated, search
#             results, agreement gate information, and each revision step done on the
#             claim.
#     """
#     original_claim = claim
#     agreement_gates = []
#     revision_steps = []

#     # Generate questions for the claim
#     questions = question_generation.run_rarr_question_generation(
#         claim=claim,
#         context=context,
#         model=model,
#         prompt=rarr_prompts.CONTEXTUAL_QGEN_PROMPT
#         if context
#         else rarr_prompts.QGEN_PROMPT,
#         temperature=temperature_qgen,
#         num_rounds=num_rounds_qgen,
#     )

#     if not questions:
#             logging.warning("No questions generated. Returning original claim.")
#             return {
#                 "context": context,
#                 "text": original_claim,
#                 "questions": [],
#                 "evidences_for_questions": [],
#                 "revisions": [],
#                 "selected_evidences": []
#             }

#     # Run search on generated question for the claim
#     if hallucinate_evidence:
#         raise_hallucinate_evidence_warning()
#         evidences_for_questions = [
#             [
#                 hallucination.run_evidence_hallucination(
#                     query=query,
#                     model=model,
#                     prompt=hallucination_prompts.EVIDENCE_HALLUCINATION,
#                 )
#             ]
#             for query in questions
#         ]
#     else:
#         evidences_for_questions = [
#             search.run_search(
#                 query=query,
#                 max_search_results_per_query=max_search_results_per_query,
#                 max_sentences_per_passage=max_sentences_per_passage,
#                 sliding_distance=sliding_distance,
#                 max_passages_per_search_result_to_return=max_passages_per_search_result,
#             )
#             for query in questions
#         ]

#     # Flatten the evidences per question into a single list.
#     used_evidences = [
#         e
#         for cur_evids in evidences_for_questions
#         for e in cur_evids[:max_evidences_per_question]
#     ]

#     # Iterative editing over each evidence
#     revision_steps = []
#     for evid in used_evidences:
#         # Run the agreement gate on the current (claim, context, query, evidence) tuple
#         gate = agreement_gate.run_agreement_gate(
#             claim=claim,
#             context=context,
#             query=evid["query"],
#             evidence=evid["text"],
#             model=model,
#             prompt=rarr_prompts.CONTEXTUAL_AGREEMENT_GATE_PROMPT
#             if context
#             else rarr_prompts.AGREEMENT_GATE_PROMPT,
#         )
#         agreement_gates.append(gate)

#         # Run the editor gate if the agreement gate is open
#         if gate["is_open"]:
#             edited_claim = editor.run_rarr_editor(
#                 claim=claim,
#                 context=context,
#                 query=evid["query"],
#                 evidence=evid["text"],
#                 model=model,
#                 prompt=rarr_prompts.CONTEXTUAL_EDITOR_PROMPT
#                 if context
#                 else rarr_prompts.EDITOR_PROMPT,
#             )["text"]

#             # Don't keep the edit if the editor makes a huge change
#             if Levenshtein.distance(claim, edited_claim) / len(claim) <= max_edit_ratio:
#                 claim = edited_claim

#         revision_steps.append({"text": claim})

#     result = {
#         "context": context,
#         "text": original_claim,
#         "questions": questions,
#         "evidences_for_questions": evidences_for_questions,
#         "revisions": [
#             {
#                 "original_text": original_claim,
#                 "revised_text": revision_steps[-1]["text"],
#                 "evidences": used_evidences,
#                 "agreement_gates": agreement_gates,
#                 "revision_steps": revision_steps,
#             }
#         ],
#     }
#     selected_evidences = evidence_selection.select_evidences(result)
#     result["selected_evidences"] = selected_evidences
#     return result


# def get_args() -> argparse.Namespace:
#     """Gets command line arguments."""
#     parser = argparse.ArgumentParser()
#     parser.add_argument(
#         "--input_file",
#         type=str,
#         required=True,
#         help="JSONLines file of claims to run RARR on.",
#     )
#     parser.add_argument(
#         "--output_file",
#         type=str,
#         required=True,
#         help="JSONLines file to write revisions to.",
#     )
#     parser.add_argument(
#         "--claim_field",
#         default="model_outputs_explanation",
#         type=str,
#         help="Field of the JSONL file to run the claim editing on.",
#     )
#     parser.add_argument(
#         "--context_field",
#         default=None,
#         type=str,
#         help="Field of the JSONL file to grab the context.",
#     )
#     parser.add_argument(
#         "--model",
#         default=" gpt-3.5-turbo",
#         type=str,
#         help="OpenAI GPT-3 model to use.",
#     )
#     parser.add_argument(
#         "--temperature_qgen",
#         default=0.7,
#         type=float,
#         help="Sampling temperature to use for query generation.",
#     )
#     parser.add_argument(
#         "--num_rounds_qgen",
#         default=3,
#         type=int,
#         help="Number of times to re-sample queries for a claim.",
#     )
#     parser.add_argument(
#         "--hallucinate_evidence",
#         action="store_true",
#         help="If this flag is set, we hallucinate evidence instead of retrieving it. "
#         "This flag should NEVER be set when trying to improve attribution as evidence  "
#         "may be inaccurate and is only provided to quickly experiment with repository "
#         "setting up the search API first.",
#     )
#     parser.add_argument(
#         "--max_search_results_per_query",
#         default=5,
#         type=int,
#         help="Maximum number of search results we get per query.",
#     )
#     parser.add_argument(
#         "--max_sentences_per_passage",
#         default=4,
#         type=int,
#         help="Maximum number of sentences per evidence passage.",
#     )
#     parser.add_argument(
#         "--sliding_distance",
#         default=1,
#         type=int,
#         help="Sliding window distance for extracting passages from a search result.",
#     )
#     parser.add_argument(
#         "--max_passages_per_search_result",
#         default=1,
#         type=int,
#         help="Maximum number of passages to return for each search result. A passage"
#         " ranker is applied to get the top passages per query.",
#     )
#     parser.add_argument(
#         "--max_evidences_per_question",
#         default=1,
#         type=int,
#         help="Maximum number of evidences to consider per question.",
#     )
#     parser.add_argument(
#         "--max_edit_ratio",
#         default=100,
#         type=float,
#         help="Maximum edit ratio between claim and edit for each round.",
#     )
#     parser.add_argument(
#         "--resume",
#         action="store_true",
#         help="Resumes the editing process if broken by loading the output file.",
#     )
#     args = parser.parse_args()

#     # Write all args to file
#     with open(args.output_file + "_args", "w", encoding="utf-8") as writer:
#         json.dump(args.__dict__, writer, indent=4)
#     return args


# def main() -> None:
#     """Loads a RARR evaluation set and runs GPT-3 RARR editing."""
#     args = get_args()

#     # Load the finished results by mapping from the claim name to the results.
#     if args.resume and os.path.exists(args.output_file):
#         print(f"Resuming with results from {args.output_file}")
#         finished_results = {
#             l["input_info"][args.claim_field]: l["result"]
#             for l in jsonlines.open(args.output_file)
#         }
#         print(f"Found {len(finished_results)} finished lines.")
#     else:
#         finished_results = None

#     with open(args.output_file, "w", encoding="utf-8") as writer:
#         lines = list(jsonlines.open(args.input_file))
#         for line in tqdm.tqdm(lines):
#             claim = line["input_info"][args.claim_field]
#             if args.context_field:
#                 context = line["input_info"][args.context_field]
#                 context = " ".join(context.split("\n"))
#             else:
#                 context = None

#             # Search for finished result
#             if finished_results and claim in finished_results:
#                 line["result"] = finished_results[claim]
#             else:
#                 line["result"] = run_editor_one_instance(
#                     model=args.model,
#                     claim=claim,
#                     context=context,
#                     temperature_qgen=args.temperature_qgen,
#                     num_rounds_qgen=args.num_rounds_qgen,
#                     max_search_results_per_query=args.max_search_results_per_query,
#                     max_sentences_per_passage=args.max_sentences_per_passage,
#                     sliding_distance=args.sliding_distance,
#                     max_passages_per_search_result=args.max_passages_per_search_result,
#                     max_evidences_per_question=args.max_evidences_per_question,
#                     max_edit_ratio=args.max_edit_ratio,
#                     hallucinate_evidence=args.hallucinate_evidence,
#                 )
#             writer.write(json.dumps(line, ensure_ascii=False) + "\n")


# if __name__ == "__main__":
#     main()
